#!/usr/bin/env python3
import sys, json, base64, io, copy
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def mk_font(bold=False, size=9, color='FF000000'):
    return Font(name='Arial', bold=bold, size=size, color=color)

def mk_fill(cor):
    return PatternFill('solid', fgColor=cor)

def mk_border(medio=False):
    if medio:
        s = Side(style='medium', color='FF8B0000')
        return Border(top=s, bottom=s)
    s = Side(style='thin', color='FFD6C4A8')
    return Border(left=s, right=s, top=s, bottom=s)

def aplicar_estilo(cell, font, fill, border, align):
    cell.font = font; cell.fill = fill; cell.border = border; cell.alignment = align

def gerar_aba(wb, nome_aba, dados):
    ws = wb.create_sheet(nome_aba)
    linhas = dados.get('linhas', [])
    meses  = dados.get('meses', [])
    ncols  = len(meses) + 2

    CORES = {
        'titulo': 'FF8B0000', 'header': 'FF5C1A1A', 'secao': 'FFF5EFE6',
        'grupo': 'FFFAF5F0', 'subtotal': 'FFECE4D8', 'branco': 'FFFFFFFF',
        'pos': 'FFD1FAE5', 'neg': 'FFFEE2E2', 'azul': 'FF1E3A5F',
        'amarelo': 'FFFEF9C3', 'cinza': 'FFF7F4F0'
    }

    # Linha 1: Título
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(1, 1, 'BOM BEEF VALINHOS — DEMONSTRATIVO DE RESULTADO')
    aplicar_estilo(c, mk_font(True,14,'FFFFFFFF'), mk_fill(CORES['titulo']),
                   mk_border(), Alignment(horizontal='center', vertical='center'))
    ws.row_dimensions[1].height = 28

    # Linha 2: Subtítulo
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c = ws.cell(2, 1, nome_aba)
    aplicar_estilo(c, mk_font(False,10,'FFFFFFFF'), mk_fill(CORES['titulo']),
                   mk_border(), Alignment(horizontal='center', vertical='center'))
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 4

    # Linha 4: Cabeçalho de meses
    c = ws.cell(4, 1, 'GRUPO / CATEGORIA')
    aplicar_estilo(c, mk_font(True,9,'FFFFFFFF'), mk_fill(CORES['header']),
                   mk_border(), Alignment(horizontal='left', vertical='center'))
    ws.row_dimensions[4].height = 22
    for i, mes in enumerate(meses):
        c = ws.cell(4, i+2, mes)
        aplicar_estilo(c, mk_font(True,9,'FFFFFFFF'), mk_fill(CORES['header']),
                       mk_border(), Alignment(horizontal='right', vertical='center'))
    c = ws.cell(4, ncols, 'TOTAL')
    aplicar_estilo(c, mk_font(True,9,'FFFFFFFF'), mk_fill(CORES['header']),
                   mk_border(), Alignment(horizontal='right', vertical='center'))

    BRL_FMT = 'R$ #,##0.00;[RED]-R$ #,##0.00'

    # Dados
    for ri, ln in enumerate(linhas, 5):
        tp  = ln.get('tipo', 'cat')
        lbl = ln.get('label', '')
        vals = ln.get('values', [])
        tot  = ln.get('total', 0)
        ws.row_dimensions[ri].height = 4 if tp == 'blank' else 16
        if tp == 'blank': continue

        # Definir estilos por tipo
        if tp == 'section':
            f = mk_font(True,10,'FF3D0000'); fl = mk_fill(CORES['secao']); b = mk_border(True)
        elif tp == 'group':
            f = mk_font(True,9,'FF5C1A1A'); fl = mk_fill(CORES['grupo']); b = mk_border()
        elif tp == 'subtotal':
            f = mk_font(True,9,'FF1A1A1A'); fl = mk_fill(CORES['subtotal']); b = mk_border()
        elif tp == 'resultado_verde':
            pos = (tot or 0) >= 0
            f = mk_font(True,11,'FF0D5C2E' if pos else 'FF7F1D1D')
            fl = mk_fill(CORES['pos'] if pos else CORES['neg']); b = mk_border(True)
        elif tp == 'resultado_final':
            pos = (tot or 0) >= 0
            f = mk_font(True,12,'FF0D5C2E' if pos else 'FF7F1D1D')
            fl = mk_fill(CORES['pos'] if pos else CORES['neg']); b = mk_border(True)
        elif tp == 'resultado_azul':
            f = mk_font(True,10,'FFFFFFFF'); fl = mk_fill(CORES['azul']); b = mk_border(True)
        elif tp == 'resultado_amarelo':
            f = mk_font(True,10,'FF4A3000'); fl = mk_fill(CORES['amarelo']); b = mk_border(True)
        elif tp == 'subtotal_small':
            f = mk_font(True,8,'FF555555'); fl = mk_fill(CORES['cinza']); b = mk_border()
        else:  # cat
            indent = 1 if '·' in lbl else 0
            f = mk_font(False,9,'FF444444'); fl = mk_fill(CORES['branco']); b = mk_border()

        indent = 2 if tp == 'cat' else 0
        c = ws.cell(ri, 1, lbl)
        aplicar_estilo(c, f, fl, b, Alignment(horizontal='left', vertical='center', indent=indent))

        for i, v in enumerate(vals):
            cv = ws.cell(ri, i+2)
            if v is not None and v != 0:
                cv.value = v
                cv.number_format = BRL_FMT
            aplicar_estilo(cv, mk_font(f.bold, f.size, f.color.rgb), fl, mk_border(),
                           Alignment(horizontal='right', vertical='center'))

        ct = ws.cell(ri, ncols)
        if tot is not None and tot != 0:
            ct.value = tot
            ct.number_format = BRL_FMT
        aplicar_estilo(ct, mk_font(True, f.size, f.color.rgb), fl, mk_border(),
                       Alignment(horizontal='right', vertical='center'))

    # Larguras
    ws.column_dimensions['A'].width = 44
    for i in range(len(meses)):
        ws.column_dimensions[get_column_letter(i+2)].width = 13
    ws.column_dimensions[get_column_letter(ncols)].width = 15
    ws.freeze_panes = ws.cell(5, 2)

if __name__ == '__main__':
    dados = json.loads(sys.stdin.read())
    wb = Workbook(); wb.remove(wb.active)
    for aba in dados.get('abas', []):
        gerar_aba(wb, aba['nome'], aba)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    print(base64.b64encode(buf.read()).decode())
